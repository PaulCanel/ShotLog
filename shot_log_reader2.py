import os
import re
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Based on original version. :contentReference[oaicite:0]{index=0}


# ==========================
# Parsing du log
# ==========================

class LogShotAnalyzer:
    def __init__(self):
        # Liste de tous les shots (dans l'ordre)
        self.shots = []  # each: dict
        # Shots encore "ouverts" pendant le parsing, indexés par (date_str, shot_number)
        self.open_shots = {}
        # Caméras attendues courantes (Updated expected cameras)
        self.current_expected = set()
        # Pour mémoriser globalement toutes les caméras vues dans expected_cameras
        self.all_expected_cameras = set()

    def parse_log_file(self, path):
        self.shots = []
        self.open_shots = {}
        self.current_expected = set()
        self.all_expected_cameras = set()

        # Regex de base
        re_updated_expected = re.compile(
            r"Updated expected cameras \(used diagnostics\): \[(.*)\]"
        )
        re_new_shot = re.compile(
            r"\*\*\* New shot detected: date=(\d{8}), shot=(\d+), camera=([A-Za-z0-9_]+), ref_time=(\d{2}:\d{2}:\d{2}) \*\*\*"
        )
        re_shot_acquired_missing_expected = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired.*?expected=\[(.*)\].*missing cameras: \[(.*)\]"
        )
        re_shot_acquired_missing = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired.*missing cameras: \[(.*)\]"
        )
        re_shot_acquired_ok_expected = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired successfully, expected=\[(.*)\], all cameras present\."
        )
        re_shot_acquired_ok = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired successfully, all cameras present\."
        )
        re_trigger_assigned = re.compile(
            r"Trigger .* assigned to existing shot (\d+).*camera ([A-Za-z0-9_]+)\)"
        )
        re_clean_copy = re.compile(
            r"CLEAN copy: .*?-> (.*)"
        )
        # Ligne timing:
        # Shot 001 (20251127) timing: trigger_cam=Lanex5, trigger_time=2025-11-27 13:33:16, ...
        re_timing = re.compile(
            r"Shot\s+(\d+)\s+\((\d{8})\)\s+timing:\s+"
            r"trigger_cam=([^,]+),\s*"
            r"trigger_time=([^,]+),\s*"
            r"min_mtime=([^,]+),\s*"
            r"max_mtime=([^,]+),\s*"
            r"first_camera=([^,]+),\s*"
            r"last_camera=([^\s,]+)"
        )

        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.rstrip("\n")

                # 1) Mise à jour des caméras attendues
                m = re_updated_expected.search(line)
                if m:
                    cam_list_text = m.group(1)
                    cams = self._parse_list_of_names(cam_list_text)
                    self.current_expected = set(cams)
                    self.all_expected_cameras.update(cams)
                    continue

                # 2) Nouveau shot détecté
                m = re_new_shot.search(line)
                if m:
                    date_str = m.group(1)
                    shot_idx = int(m.group(2))
                    cam = m.group(3)
                    ref_time_str = m.group(4)  # HH:MM:SS (ref mtime du trigger)

                    shot = {
                        "date": date_str,
                        "shot_number": shot_idx,
                        "trigger_cams": {cam},  # caméras ayant servi de trigger
                        "expected_cams": set(),   # remplie à la fin
                        "missing_cams": set(),
                        "image_times": [],        # (de CLEAN copy si besoin)
                        # Champs timing qui viennent de la ligne "timing:"
                        "trigger_camera": None,
                        "trigger_time": None,
                        "min_time": None,
                        "max_time": None,
                        "first_camera": None,
                        "last_camera": None,
                    }
                    self.shots.append(shot)
                    self.open_shots[(date_str, shot_idx)] = shot
                    continue

                # 3) Trigger assigné à un shot existant (trigger secondaire)
                m = re_trigger_assigned.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    cam = m.group(2)
                    shot = self._find_open_shot_by_index(shot_idx)
                    if shot is not None:
                        shot["trigger_cams"].add(cam)
                    continue

                # 4) CLEAN copy => on récupère chemin destination, parse le nom
                m = re_clean_copy.search(line)
                if m:
                    dest_path = m.group(1).strip()
                    filename = os.path.basename(dest_path)
                    # Forme: Cam_YYYYMMDD_HHMMSS_shotNNN.tif
                    name_match = re.match(
                        r"([A-Za-z0-9_]+)_(\d{8})_(\d{6})_shot(\d+)", filename
                    )
                    if name_match:
                        cam = name_match.group(1)
                        date_str = name_match.group(2)
                        time_str = name_match.group(3)
                        shot_idx = int(name_match.group(4))
                        dt = self._parse_datetime(date_str, time_str)
                        shot = self._find_open_or_recent_shot(date_str, shot_idx)
                        if shot is not None and dt is not None:
                            shot["image_times"].append(dt)
                    continue

                # 5) Shot avec missing cameras
                m = re_shot_acquired_missing_expected.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    expected_text = m.group(3)
                    missing_text = m.group(4)
                    expected_cams = set(self._parse_list_of_names(expected_text))
                    missing_cams = set(self._parse_list_of_names(missing_text))

                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = expected_cams
                        shot["missing_cams"] = missing_cams
                        self.all_expected_cameras.update(expected_cams)
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                m = re_shot_acquired_missing.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    missing_text = m.group(3)
                    missing_cams = set(self._parse_list_of_names(missing_text))

                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = set(self.current_expected)
                        shot["missing_cams"] = missing_cams
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                # 6) Shot sans missing (tout OK)
                m = re_shot_acquired_ok_expected.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    expected_text = m.group(3)
                    expected_cams = set(self._parse_list_of_names(expected_text))
                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = expected_cams
                        shot["missing_cams"] = set()
                        self.all_expected_cameras.update(expected_cams)
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                m = re_shot_acquired_ok.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = set(self.current_expected)
                        shot["missing_cams"] = set()
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                # 7) Ligne timing avec trigger/min/max/first/last
                m = re_timing.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    trigger_cam = m.group(3).strip()
                    trigger_time_str = m.group(4).strip()
                    min_time_str = m.group(5).strip()
                    max_time_str = m.group(6).strip()
                    first_cam = m.group(7).strip()
                    last_cam = m.group(8).strip()

                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["trigger_camera"] = trigger_cam
                        shot["trigger_time"] = self._parse_datetime_full(trigger_time_str)
                        shot["min_time"] = self._parse_datetime_full(min_time_str)
                        shot["max_time"] = self._parse_datetime_full(max_time_str)
                        # first/last camera peuvent être "N/A"
                        shot["first_camera"] = first_cam if first_cam != "N/A" else None
                        shot["last_camera"] = last_cam if last_cam != "N/A" else None
                    continue

        # Si certains shots n'ont pas de ligne timing, on récupère au moins min/max via image_times
        for shot in self.shots:
            if shot["min_time"] is None and shot["image_times"]:
                shot["min_time"] = min(shot["image_times"])
            if shot["max_time"] is None and shot["image_times"]:
                shot["max_time"] = max(shot["image_times"])

        return self.shots

    # ---------------------------
    #  Méthodes utilitaires
    # ---------------------------

    @staticmethod
    def _parse_list_of_names(text):
        """
        Texte du style: "'Lanex1', 'Lanex2', 'DarkShadow'"
        -> ["Lanex1", "Lanex2", "DarkShadow"]
        """
        if not text.strip():
            return []
        parts = text.split(",")
        names = []
        for p in parts:
            s = p.strip()
            if s.startswith("'") or s.startswith('"'):
                s = s[1:]
            if s.endswith("'") or s.endswith('"'):
                s = s[:-1]
            if s:
                names.append(s)
        return names

    @staticmethod
    def _parse_datetime(date_str, time_str):
        try:
            return datetime.strptime(date_str + time_str, "%Y%m%d%H%M%S")
        except Exception:
            return None

    @staticmethod
    def _parse_datetime_full(dt_str):
        """
        dt_str du type '2025-11-27 13:33:16'
        """
        if dt_str is None or dt_str == "N/A":
            return None
        try:
            return datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None

    def _find_open_shot_by_index(self, shot_idx):
        # Si plusieurs dates, on cherche dans open_shots
        for (d, i), shot in self.open_shots.items():
            if i == shot_idx:
                return shot
        return None

    def _find_open_or_recent_shot(self, date_str, shot_idx):
        key = (date_str, shot_idx)
        if key in self.open_shots:
            return self.open_shots[key]
        # sinon, on prend le dernier shot dans self.shots avec cette date/numéro
        for shot in reversed(self.shots):
            if shot["date"] == date_str and shot["shot_number"] == shot_idx:
                return shot
        return None

    # ---------------------------
    #  Résumés pour l'interface
    # ---------------------------

    def compute_global_summary(self):
        if not self.shots:
            return {
                "dates": [],
                "start_time": None,
                "end_time": None,
                "total_shots": 0,
                "shots_with_missing": 0,
            }

        # Date(s) distinctes présentes dans les shots
        dates = sorted({s["date"] for s in self.shots})

        # Min / max temps global (on privilégie min_time/max_time si présents)
        all_starts = []
        all_ends = []
        for s in self.shots:
            if s["min_time"] is not None:
                all_starts.append(s["min_time"])
            if s["max_time"] is not None:
                all_ends.append(s["max_time"])
        start_time = min(all_starts) if all_starts else None
        end_time = max(all_ends) if all_ends else None

        total_shots = len(self.shots)
        shots_with_missing = sum(1 for s in self.shots if s["missing_cams"])

        return {
            "dates": dates,
            "start_time": start_time,
            "end_time": end_time,
            "total_shots": total_shots,
            "shots_with_missing": shots_with_missing,
        }

    def compute_camera_summary(self):
        # Caméras vues dans expected_cams
        cams = sorted({c for s in self.shots for c in s["expected_cams"]})
        summary = []
        for cam in cams:
            used_count = sum(1 for s in self.shots if cam in s["expected_cams"])
            missing_count = sum(1 for s in self.shots if cam in s["missing_cams"])
            summary.append(
                {"camera": cam, "shots_used": used_count, "shots_missing": missing_count}
            )
        return summary


# ==========================
# Watchdog handler
# ==========================

class LogFileEventHandler(FileSystemEventHandler):
    def __init__(self, gui, path):
        super().__init__()
        self.gui = gui
        self._path = os.path.abspath(path)

    def on_modified(self, event):
        if os.path.abspath(event.src_path) == self._path:
            self.gui.root.after(0, self.gui.reload_current_log_from_watcher)

    def on_moved(self, event):
        # Si le fichier est renommé ou déplacé, on ne fait rien pour l'instant
        pass


# ==========================
# GUI Tkinter
# ==========================

class LogAnalyzerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Shot Log Analyzer (XLSX)")

        self.analyzer = LogShotAnalyzer()
        self.current_log_path = None
        self.shots_data = []
        self.camera_summary = []
        self.global_summary = {}

        self.observer = None
        self._watched_handler = None

        self._build_gui()

    def _build_gui(self):
        # Barre d'outils en haut
        toolbar = ttk.Frame(self.root)
        toolbar.pack(fill="x", padx=5, pady=5)

        btn_open = ttk.Button(toolbar, text="Open log...", command=self.load_log)
        btn_open.pack(side="left", padx=5)

        self.btn_save = ttk.Button(toolbar, text="Save XLSX", command=self.save_xlsx, state="disabled")
        self.btn_save.pack(side="left", padx=5)

        # Label pour afficher le chemin du log
        self.lbl_logpath = ttk.Label(toolbar, text="No log loaded")
        self.lbl_logpath.pack(side="left", padx=10)

        # Résumé global
        frm_global = ttk.LabelFrame(self.root, text="Global summary")
        frm_global.pack(fill="x", padx=5, pady=5)

        self.lbl_global = ttk.Label(frm_global, text="No data")
        self.lbl_global.pack(anchor="w", padx=5, pady=5)

        # Résumé par caméra
        frm_cam = ttk.LabelFrame(self.root, text="Per-camera summary")
        frm_cam.pack(fill="x", padx=5, pady=5)

        columns_cam = ("camera", "shots_used", "shots_missing")
        self.tree_cam = ttk.Treeview(frm_cam, columns=columns_cam, show="headings", height=5)
        for col in columns_cam:
            self.tree_cam.heading(col, text=col)
            self.tree_cam.column(col, width=140, anchor="center")
        self.tree_cam.pack(side="left", fill="x", expand=True)

        scrollbar_cam = ttk.Scrollbar(frm_cam, orient="vertical", command=self.tree_cam.yview)
        self.tree_cam.configure(yscrollcommand=scrollbar_cam.set)
        scrollbar_cam.pack(side="right", fill="y")

        # Tableur par shot
        frm_shots = ttk.LabelFrame(self.root, text="Shots")
        frm_shots.pack(fill="both", expand=True, padx=5, pady=5)

        # Colonnes demandées :
        # shot number, missing, expected, trigger time, min time, max time,
        # missing list, trigger camera, first camera, last camera,
        # expected list, trigger list
        columns_shot = (
            "shot_number",
            "missing_count",
            "expected_count",
            "trigger_time",
            "min_time",
            "max_time",
            "missing_cams",
            "trigger_camera",
            "first_camera",
            "last_camera",
            "expected_cams",
            "trigger_cams",
        )
        headers = {
            "shot_number": "Shot #",
            "missing_count": "# Missing",
            "expected_count": "# Expected",
            "trigger_time": "Trigger time",
            "min_time": "Min time",
            "max_time": "Max time",
            "missing_cams": "Missing cameras",
            "trigger_camera": "Trigger camera",
            "first_camera": "First camera",
            "last_camera": "Last camera",
            "expected_cams": "Expected cameras",
            "trigger_cams": "Trigger cameras",
        }

        self.tree_shot = ttk.Treeview(frm_shots, columns=columns_shot, show="headings")
        for col in columns_shot:
            self.tree_shot.heading(col, text=headers[col])
            if col in ("missing_cams", "expected_cams", "trigger_cams"):
                width = 260
            else:
                width = 120
            self.tree_shot.column(col, width=width, anchor="center")

        self.tree_shot.pack(side="left", fill="both", expand=True)

        scrollbar_shot = ttk.Scrollbar(frm_shots, orient="vertical", command=self.tree_shot.yview)
        self.tree_shot.configure(yscrollcommand=scrollbar_shot.set)
        scrollbar_shot.pack(side="right", fill="y")

        # Styles / tags pour lignes OK / MISSING (on joue sur la couleur)
        self.tree_shot.tag_configure("ok", background="green", foreground="white")
        self.tree_shot.tag_configure("missing", background="red", foreground="white")

        # Résumé caméras : tags aussi (on colore toute la ligne, limitation Treeview)
        self.tree_cam.tag_configure("cam_ok", background="green", foreground="white")
        self.tree_cam.tag_configure("cam_missing", background="red", foreground="white")

        # Internal logs
        frm_log = ttk.LabelFrame(self.root, text="Internal messages")
        frm_log.pack(fill="both", expand=False, padx=5, pady=5)

        self.txt_log = scrolledtext.ScrolledText(frm_log, height=8)
        self.txt_log.pack(fill="both", expand=True)
        self._log_message("Ready. Open a log file to analyze it.")

    # ---------------------------
    # Watchdog control
    # ---------------------------

    def start_watching(self):
        self.stop_watching()
        if not self.current_log_path:
            return
        directory = os.path.dirname(self.current_log_path) or "."
        handler = LogFileEventHandler(self, self.current_log_path)
        observer = Observer()
        observer.schedule(handler, directory, recursive=False)
        observer.daemon = True
        observer.start()
        self.observer = observer
        self._watched_handler = handler
        self._log_message(f"Watching {self.current_log_path} for changes...")

    def stop_watching(self):
        if self.observer is not None:
            self.observer.stop()
            self.observer = None
            self._watched_handler = None
            self._log_message("Stopped watching log file.")

    def reload_current_log_from_watcher(self):
        if not self.current_log_path:
            return
        try:
            shots = self.analyzer.parse_log_file(self.current_log_path)
        except Exception as e:
            self._log_message(f"Auto-reload failed: {e}")
            return

        self.shots_data = shots
        self.camera_summary = self.analyzer.compute_camera_summary()
        self.global_summary = self.analyzer.compute_global_summary()

        self._log_message(f"Detected change, reloaded log: {self.current_log_path}")
        self._refresh_global_summary()
        self._refresh_camera_table()
        self._refresh_shots_table()

        if self.shots_data:
            self.btn_save.configure(state="normal")
        else:
            self.btn_save.configure(state="disabled")

    # ---------------------------
    # Actions
    # ---------------------------

    def _log_message(self, msg):
        self.txt_log.insert("end", msg + "\n")
        self.txt_log.see("end")

    def load_log(self):
        path = filedialog.askopenfilename(
            title="Select a log file",
            filetypes=[("Text files", "*.txt;*.log"), ("All files", "*.*")]
        )
        if not path:
            return

        try:
            shots = self.analyzer.parse_log_file(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse log: {e}")
            return

        self.current_log_path = path
        self.shots_data = shots
        self.camera_summary = self.analyzer.compute_camera_summary()
        self.global_summary = self.analyzer.compute_global_summary()

        self.lbl_logpath.configure(text=os.path.basename(path))
        self._log_message(f"Loaded log: {path}")
        self._refresh_global_summary()
        self._refresh_camera_table()
        self._refresh_shots_table()

        if self.shots_data:
            self.btn_save.configure(state="normal")
        else:
            self.btn_save.configure(state="disabled")

        self.start_watching()

    def _refresh_global_summary(self):
        gs = self.global_summary
        if not gs or gs["total_shots"] == 0:
            self.lbl_global.configure(text="No shots found in log.")
            return

        dates = gs["dates"]
        if len(dates) == 1:
            date_txt = dates[0]
        else:
            date_txt = ", ".join(dates)

        def fmt_time(dt):
            return dt.strftime("%H:%M:%S") if dt else "-"

        txt = (
            f"Date(s): {date_txt} | "
            f"Start: {fmt_time(gs['start_time'])} | "
            f"End: {fmt_time(gs['end_time'])} | "
            f"Total shots: {gs['total_shots']} | "
            f"Shots with missing cameras: {gs['shots_with_missing']}"
        )
        self.lbl_global.configure(text=txt)

    def _refresh_camera_table(self):
        self.tree_cam.delete(*self.tree_cam.get_children())
        for row in self.camera_summary:
            cam = row["camera"]
            used = row["shots_used"]
            missing = row["shots_missing"]
            tag = "cam_ok" if missing == 0 else "cam_missing"
            self.tree_cam.insert(
                "", "end",
                values=(cam, used, missing),
                tags=(tag,)
            )

    def _refresh_shots_table(self):
        self.tree_shot.delete(*self.tree_shot.get_children())

        def fmt_time(dt):
            return dt.strftime("%H:%M:%S") if dt else ""

        for shot in self.shots_data:
            missing_cams = sorted(shot["missing_cams"]) if shot["missing_cams"] else []
            expected_cams = sorted(shot["expected_cams"]) if shot["expected_cams"] else []
            trigger_cams = sorted(shot["trigger_cams"]) if shot["trigger_cams"] else []

            # Listes au format [a, b, c] pour rester dans une seule cellule en Excel
            missing_str = "[" + ", ".join(missing_cams) + "]"
            expected_str = "[" + ", ".join(expected_cams) + "]"
            trigger_str = "[" + ", ".join(trigger_cams) + "]"

            missing_count = len(missing_cams)
            expected_count = len(expected_cams)

            trigger_cam = shot.get("trigger_camera") or ""
            first_cam = shot.get("first_camera") or ""
            last_cam = shot.get("last_camera") or ""

            values = (
                f"{shot['shot_number']:03d}",
                missing_count,
                expected_count,
                fmt_time(shot["trigger_time"]),
                fmt_time(shot["min_time"]),
                fmt_time(shot["max_time"]),
                missing_str,
                trigger_cam,
                first_cam,
                last_cam,
                expected_str,
                trigger_str,
            )

            # Tag selon missing (toute la ligne en vert/rouge, texte blanc)
            tag = "ok" if missing_count == 0 else "missing"
            self.tree_shot.insert("", "end", values=values, tags=(tag,))

    def save_xlsx(self):
        if not self.shots_data:
            messagebox.showwarning("Warning", "No shot data to save.")
            return

        path = filedialog.asksaveasfilename(
            title="Save XLSX",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return

        try:
            wb = Workbook()

            # Styles Excel
            ok_fill = PatternFill("solid", fgColor="00AA00")     # vert
            miss_fill = PatternFill("solid", fgColor="CC0000")   # rouge
            white_font = Font(color="FFFFFF", bold=True)
            black_font = Font(color="000000", bold=False)

            # 1) Summary
            ws_summary = wb.active
            ws_summary.title = "Summary"

            gs = self.global_summary

            def fmt_time(dt):
                return dt.strftime("%H:%M:%S") if dt else ""

            if gs and gs["total_shots"] > 0:
                dates = gs["dates"]
                if len(dates) == 1:
                    date_txt = dates[0]
                else:
                    date_txt = ", ".join(dates)

                ws_summary["A1"] = "Dates"
                ws_summary["B1"] = date_txt
                ws_summary["A2"] = "Start"
                ws_summary["B2"] = fmt_time(gs["start_time"])
                ws_summary["A3"] = "End"
                ws_summary["B3"] = fmt_time(gs["end_time"])
                ws_summary["A4"] = "Total shots"
                ws_summary["B4"] = gs["total_shots"]
                ws_summary["A5"] = "Shots with missing"
                ws_summary["B5"] = gs["shots_with_missing"]
            else:
                ws_summary["A1"] = "No shots in log."

            # 2) Cameras
            ws_cam = wb.create_sheet("Cameras")
            ws_cam.append(["Camera", "Shots requested", "Shots missing for this camera"])

            for row in self.camera_summary:
                cam = row["camera"]
                used = row["shots_used"]
                missing = row["shots_missing"]
                excel_row = ws_cam.max_row + 1
                ws_cam.append([cam, used, missing])

                # Colorer uniquement la première cellule dans Excel
                cell = ws_cam[f"A{excel_row}"]
                if missing == 0:
                    cell.fill = ok_fill
                    cell.font = white_font
                else:
                    cell.fill = miss_fill
                    cell.font = white_font

            # 3) Shots
            ws_shots = wb.create_sheet("Shots")
            ws_shots.append([
                "Shot #",
                "# Missing",
                "# Expected",
                "Trigger time",
                "Min time",
                "Max time",
                "Missing cameras",
                "Trigger camera",
                "First camera",
                "Last camera",
                "Expected cameras",
                "Trigger cameras",
            ])

            for shot in self.shots_data:
                missing_cams = sorted(shot["missing_cams"]) if shot["missing_cams"] else []
                expected_cams = sorted(shot["expected_cams"]) if shot["expected_cams"] else []
                trigger_cams = sorted(shot["trigger_cams"]) if shot["trigger_cams"] else []

                missing_str = "[" + ", ".join(missing_cams) + "]"
                expected_str = "[" + ", ".join(expected_cams) + "]"
                trigger_str = "[" + ", ".join(trigger_cams) + "]"

                missing_count = len(missing_cams)
                expected_count = len(expected_cams)

                trigger_cam = shot.get("trigger_camera") or ""
                first_cam = shot.get("first_camera") or ""
                last_cam = shot.get("last_camera") or ""

                row_values = [
                    f"{shot['shot_number']:03d}",
                    missing_count,
                    expected_count,
                    fmt_time(shot["trigger_time"]),
                    fmt_time(shot["min_time"]),
                    fmt_time(shot["max_time"]),
                    missing_str,
                    trigger_cam,
                    first_cam,
                    last_cam,
                    expected_str,
                    trigger_str,
                ]

                ws_shots.append(row_values)
                excel_row = ws_shots.max_row

                # Colorer toute la ligne en fonction de missing_count
                if missing_count == 0:
                    fill = ok_fill
                else:
                    fill = miss_fill

                for cell in ws_shots[excel_row]:
                    cell.fill = fill
                    cell.font = white_font

            # Largeurs de colonnes approximatives
            for col in ws_shots.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                ws_shots.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))

            wb.save(path)
            messagebox.showinfo("Saved", f"XLSX saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save XLSX: {e}")


# ==========================
# main
# ==========================

def main():
    root = tk.Tk()
    root.geometry("1400x850")
    app = LogAnalyzerGUI(root)

    def on_close():
        app.stop_watching()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
 
