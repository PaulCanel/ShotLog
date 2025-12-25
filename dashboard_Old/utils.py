"""Generic helpers for the Streamlit dashboard."""
from __future__ import annotations

import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from data_models import CombinedAlignment, DisplayRow, ParsedLog, ParsedManual, ParsedMotor

GREEN_BG = "#006400"
BLUE_BG = "#00008B"
RED_BG = "#8B0000"
ORANGE_BG = "#FF8C00"
TEXT_DEFAULT = "white"
TEXT_WARNING = "#FFFF00"


def format_datetime(dt: datetime | None) -> str:
    return dt.strftime("%H:%M:%S") if dt else "-"


def format_ratio(ratio: float) -> str:
    return f"{ratio*100:.1f}%"


def seconds_to_clock(value: float | None) -> str:
    if value is None:
        return "00:00:00"
    total_seconds = int(round(value))
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def file_signature(path: str | os.PathLike[str]) -> tuple[str, float] | None:
    try:
        stat = os.stat(path)
    except FileNotFoundError:
        return None
    return str(path), stat.st_mtime


def ensure_exports_dir() -> Path:
    exports = Path("exports")
    exports.mkdir(exist_ok=True)
    return exports


def export_to_excel(
    log_data: ParsedLog,
    manual: ParsedManual,
    motor: ParsedMotor,
    alignment: CombinedAlignment,
    dest_path: Path,
) -> Path:
    """Replicate the legacy Excel export with the new datamodels."""

    wb = Workbook()

    ws_logs = wb.active
    ws_logs.title = "Logs"
    log_headers = [
        "Shot #",
        "# Missing",
        "# Expected",
        "Trigger time",
        "Min time",
        "Max time",
        "Missing cameras",
        "Trigger camera",
        "First camera",
        "Last camera",
        "Expected cameras",
        "Trigger cameras",
    ]
    ws_logs.append(log_headers)

    for shot in log_data.shots:
        missing_cams = sorted(shot.missing_cams)
        expected_cams = sorted(shot.expected_cams)
        trigger_cams = sorted(shot.trigger_cams)

        values = [
            f"{shot.shot_number:03d}",
            len(missing_cams),
            len(expected_cams),
            format_datetime(shot.trigger_time),
            format_datetime(shot.min_time),
            format_datetime(shot.max_time),
            "[" + ", ".join(missing_cams) + "]",
            shot.trigger_camera or "",
            shot.first_camera or "",
            shot.last_camera or "",
            "[" + ", ".join(expected_cams) + "]",
            "[" + ", ".join(trigger_cams) + "]",
        ]
        row_obj = DisplayRow(
            key=(shot.shot_number, format_datetime(shot.trigger_time) if shot.trigger_time else ""),
            values=[str(v) for v in values],
            bg="green" if not shot.missing_cams else "red",
            incomplete=bool(shot.missing_cams),
            yellow_text=False,
        )
        ws_logs.append(values)
        _apply_excel_styles(ws_logs, ws_logs.max_row, row_obj, alignment.yellow_keys)

    ws_cam = wb.create_sheet("Per_camera_summary")
    ws_cam.append(["Camera", "Shots requested", "Shots missing for this camera"])
    for row in log_data.per_camera_summary:
        ws_cam.append([row.camera, row.shots_used, row.shots_missing])
        excel_row = ws_cam.max_row
        fill_color = GREEN_BG.lstrip("#") if row.shots_missing == 0 else RED_BG.lstrip("#")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        font = Font(color=_to_argb(TEXT_DEFAULT))
        for cell in ws_cam[excel_row]:
            cell.fill = fill
            cell.font = font

    ws_manual = wb.create_sheet("Manual_Params")
    ws_motor = wb.create_sheet("Motor_Params")
    for ws, rows, headers in (
        (ws_manual, alignment.manual_rows, manual.header),
        (ws_motor, alignment.motor_rows, motor.header),
    ):
        ws.append(headers)
        for row in rows:
            ws.append(row.values)
            _apply_excel_styles(ws, ws.max_row, row, alignment.yellow_keys)

    dest_path.parent.mkdir(exist_ok=True, parents=True)
    wb.save(dest_path)
    return dest_path


def _apply_excel_styles(ws, row_idx: int, row: DisplayRow, yellow_keys):
    color_map = {
        "green": GREEN_BG.lstrip("#"),
        "blue": BLUE_BG.lstrip("#"),
        "red": RED_BG.lstrip("#"),
        "orange": ORANGE_BG.lstrip("#"),
    }
    fill_color = color_map.get(row.bg)
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
    font_color = TEXT_WARNING.lstrip("#") if (row.key in yellow_keys or row.yellow_text) else TEXT_DEFAULT
    font = Font(color=_to_argb(font_color))
    for cell in ws[row_idx]:
        if fill:
            cell.fill = fill
        if font:
            cell.font = font


def _to_argb(color: str) -> str:
    if not color:
        return "FFFFFFFF"
    normalized = color.strip()
    if normalized.startswith("#"):
        normalized = normalized[1:]
    if normalized.lower() == "white":
        normalized = "FFFFFF"
    if len(normalized) == 8:
        return normalized.upper()
    if len(normalized) == 6:
        return f"FF{normalized.upper()}"
    return "FFFFFFFF"


__all__ = [
    "format_datetime",
    "format_ratio",
    "seconds_to_clock",
    "file_signature",
    "ensure_exports_dir",
    "export_to_excel",
]
