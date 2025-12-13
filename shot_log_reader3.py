import csv
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from statistics import mean, median

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer


GREEN_BG = "#006400"  # DarkGreen
BLUE_BG = "#00008B"  # DarkBlue
RED_BG = "#8B0000"  # DarkRed
ORANGE_BG = "#FF8C00"  # DarkOrange
TEXT_DEFAULT = "white"
TEXT_WARNING = "#FFFF00"


def to_argb(color: str) -> str:
    """Convert a color string to ARGB format required by openpyxl.

    Accepted inputs: "white", "#RRGGBB", "RRGGBB", "AARRGGBB".
    Returns upper-case "AARRGGBB", defaulting to opaque white.
    """

    if not color:
        return "FFFFFFFF"

    normalized = color.strip()
    if normalized.startswith("#"):
        normalized = normalized[1:]

    if normalized.lower() == "white":
        normalized = "FFFFFF"

    if len(normalized) == 8:
        return normalized.upper()
    if len(normalized) == 6:
        return f"FF{normalized.upper()}"

    return "FFFFFFFF"


# ==========================
# Helpers / models
# ==========================


@dataclass
class ShotViewRow:
    key: tuple[int, str]
    values: list[str]
    bg: str
    yellow_text: bool
    incomplete: bool = False


class LogShotAnalyzer:
    def __init__(self):
        self.shots: list[dict] = []
        self.open_shots: dict[tuple[str, int], dict] = {}
        self.current_expected: set[str] = set()
        self.all_expected_cameras: set[str] = set()

    def parse_log_file(self, path: Path) -> list[dict]:
        self.shots = []
        self.open_shots = {}
        self.current_expected = set()
        self.all_expected_cameras = set()

        re_updated_expected = re.compile(r"Updated expected cameras \(used diagnostics\): \[(.*)\]")
        re_new_shot = re.compile(
            r"\*\*\* New shot detected: date=(\d{8}), shot=(\d+), camera=([A-Za-z0-9_]+), ref_time=(\d{2}:\d{2}:\d{2}) \*\*\*"
        )
        re_shot_acquired_missing_expected = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired.*?expected=\[(.*)\].*missing cameras: \[(.*)\]"
        )
        re_shot_acquired_missing = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired.*missing cameras: \[(.*)\]"
        )
        re_shot_acquired_ok_expected = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired successfully, expected=\[(.*)\], all cameras present\."
        )
        re_shot_acquired_ok = re.compile(
            r"Shot (\d+) \((\d{8})\) acquired successfully, all cameras present\."
        )
        re_trigger_assigned = re.compile(r"Trigger .* assigned to existing shot (\d+).*camera ([A-Za-z0-9_]+)\)")
        re_clean_copy = re.compile(r"CLEAN copy: .*?-> (.*)")
        re_timing = re.compile(
            r"Shot\s+(\d+)\s+\((\d{8})\)\s+timing:\s+"
            r"trigger_cam=([^,]+),\s*"
            r"trigger_time=([^,]+),\s*"
            r"min_mtime=([^,]+),\s*"
            r"max_mtime=([^,]+),\s*"
            r"first_camera=([^,]+),\s*"
            r"last_camera=([^\s,]+)"
        )

        with path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.rstrip("\n")

                m = re_updated_expected.search(line)
                if m:
                    cam_list_text = m.group(1)
                    cams = self._parse_list_of_names(cam_list_text)
                    self.current_expected = set(cams)
                    self.all_expected_cameras.update(cams)
                    continue

                m = re_new_shot.search(line)
                if m:
                    date_str = m.group(1)
                    shot_idx = int(m.group(2))
                    cam = m.group(3)
                    shot = {
                        "date": date_str,
                        "shot_number": shot_idx,
                        "trigger_cams": {cam},
                        "expected_cams": set(),
                        "missing_cams": set(),
                        "image_times": [],
                        "trigger_camera": None,
                        "trigger_time": None,
                        "min_time": None,
                        "max_time": None,
                        "first_camera": None,
                        "last_camera": None,
                    }
                    self.shots.append(shot)
                    self.open_shots[(date_str, shot_idx)] = shot
                    continue

                m = re_trigger_assigned.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    cam = m.group(2)
                    shot = self._find_open_shot_by_index(shot_idx)
                    if shot is not None:
                        shot["trigger_cams"].add(cam)
                    continue

                m = re_clean_copy.search(line)
                if m:
                    dest_path = m.group(1).strip()
                    filename = os.path.basename(dest_path)
                    name_match = re.match(r"([A-Za-z0-9_]+)_(\d{8})_(\d{6})_shot(\d+)", filename)
                    if name_match:
                        cam = name_match.group(1)
                        date_str = name_match.group(2)
                        time_str = name_match.group(3)
                        shot_idx = int(name_match.group(4))
                        dt = self._parse_datetime(date_str, time_str)
                        shot = self._find_open_or_recent_shot(date_str, shot_idx)
                        if shot is not None and dt is not None:
                            shot["image_times"].append(dt)
                    continue

                m = re_shot_acquired_missing_expected.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    expected_text = m.group(3)
                    missing_text = m.group(4)
                    expected_cams = set(self._parse_list_of_names(expected_text))
                    missing_cams = set(self._parse_list_of_names(missing_text))

                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = expected_cams
                        shot["missing_cams"] = missing_cams
                        self.all_expected_cameras.update(expected_cams)
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                m = re_shot_acquired_missing.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    missing_text = m.group(3)
                    missing_cams = set(self._parse_list_of_names(missing_text))

                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = set(self.current_expected)
                        shot["missing_cams"] = missing_cams
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                m = re_shot_acquired_ok_expected.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    expected_text = m.group(3)
                    expected_cams = set(self._parse_list_of_names(expected_text))
                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = expected_cams
                        shot["missing_cams"] = set()
                        self.all_expected_cameras.update(expected_cams)
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                m = re_shot_acquired_ok.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["expected_cams"] = set(self.current_expected)
                        shot["missing_cams"] = set()
                    self.open_shots.pop((date_str, shot_idx), None)
                    continue

                m = re_timing.search(line)
                if m:
                    shot_idx = int(m.group(1))
                    date_str = m.group(2)
                    trigger_cam = m.group(3).strip()
                    trigger_time_str = m.group(4).strip()
                    min_time_str = m.group(5).strip()
                    max_time_str = m.group(6).strip()
                    first_cam = m.group(7).strip()
                    last_cam = m.group(8).strip()

                    shot = self._find_open_or_recent_shot(date_str, shot_idx)
                    if shot is not None:
                        shot["trigger_camera"] = trigger_cam
                        shot["trigger_time"] = self._parse_datetime_full(trigger_time_str)
                        shot["min_time"] = self._parse_datetime_full(min_time_str)
                        shot["max_time"] = self._parse_datetime_full(max_time_str)
                        shot["first_camera"] = first_cam if first_cam != "N/A" else None
                        shot["last_camera"] = last_cam if last_cam != "N/A" else None
                    continue

        for shot in self.shots:
            if shot["min_time"] is None and shot["image_times"]:
                shot["min_time"] = min(shot["image_times"])
            if shot["max_time"] is None and shot["image_times"]:
                shot["max_time"] = max(shot["image_times"])

        return self.shots

    @staticmethod
    def _parse_list_of_names(text: str):
        if not text.strip():
            return []
        parts = text.split(",")
        names = []
        for p in parts:
            s = p.strip()
            if s.startswith("'") or s.startswith('"'):
                s = s[1:]
            if s.endswith("'") or s.endswith('"'):
                s = s[:-1]
            if s:
                names.append(s)
        return names

    @staticmethod
    def _parse_datetime(date_str, time_str):
        try:
            return datetime.strptime(date_str + time_str, "%Y%m%d%H%M%S")
        except Exception:
            return None

    @staticmethod
    def _parse_datetime_full(dt_str):
        if dt_str is None or dt_str == "N/A":
            return None
        try:
            return datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None

    def _find_open_shot_by_index(self, shot_idx):
        for (d, i), shot in self.open_shots.items():
            if i == shot_idx:
                return shot
        return None

    def _find_open_or_recent_shot(self, date_str, shot_idx):
        key = (date_str, shot_idx)
        if key in self.open_shots:
            return self.open_shots[key]
        for shot in reversed(self.shots):
            if shot["date"] == date_str and shot["shot_number"] == shot_idx:
                return shot
        return None

    def compute_global_summary(self):
        if not self.shots:
            return {
                "dates": [],
                "start_time": None,
                "end_time": None,
                "total_shots": 0,
                "shots_with_missing": 0,
            }

        dates = sorted({s["date"] for s in self.shots})

        all_starts = []
        all_ends = []
        for s in self.shots:
            if s["min_time"] is not None:
                all_starts.append(s["min_time"])
            if s["max_time"] is not None:
                all_ends.append(s["max_time"])
        start_time = min(all_starts) if all_starts else None
        end_time = max(all_ends) if all_ends else None

        total_shots = len(self.shots)
        shots_with_missing = sum(1 for s in self.shots if s["missing_cams"])

        return {
            "dates": dates,
            "start_time": start_time,
            "end_time": end_time,
            "total_shots": total_shots,
            "shots_with_missing": shots_with_missing,
        }

    def compute_camera_summary(self):
        cams = sorted({c for s in self.shots for c in s["expected_cams"]})
        summary = []
        for cam in cams:
            used_count = sum(1 for s in self.shots if cam in s["expected_cams"])
            missing_count = sum(1 for s in self.shots if cam in s["missing_cams"])
            summary.append({"camera": cam, "shots_used": used_count, "shots_missing": missing_count})
        return summary


# ==========================
# Watchdog wrapper
# ==========================

class TargetWatcher(FileSystemEventHandler):
    def __init__(self, path: Path, callback, root: tk.Tk):
        super().__init__()
        self.path = path
        self.callback = callback
        self.root = root

    def _schedule(self):
        if self.root:
            self.root.after(0, self.callback)
        else:
            self.callback()

    def on_modified(self, event):
        if Path(event.src_path) == self.path:
            self._schedule()

    def on_created(self, event):
        if Path(event.src_path) == self.path:
            self._schedule()


# ==========================
# Plotting panel
# ==========================


class ColumnPlotPanel:
    def __init__(self, parent: tk.Widget, color: str):
        self.color = color
        self.frame = ttk.Frame(parent)
        self.frame.pack(fill="x", padx=5, pady=5)

        self.grid_var = tk.BooleanVar(value=True)
        self.bins_var = tk.IntVar(value=5)
        self.mode = "plot"

        controls = ttk.Frame(self.frame)
        controls.pack(fill="x", pady=2)

        ttk.Checkbutton(controls, text="Grid", variable=self.grid_var, command=self._toggle_grid).pack(
            side="left", padx=5
        )
        self.toggle_btn = ttk.Button(controls, text="Histogram", command=self._toggle_mode)
        self.toggle_btn.pack(side="left", padx=5)

        ttk.Label(controls, text="Bins:").pack(side="left")
        self.bins_spin = tk.Spinbox(
            controls,
            from_=1,
            to=100,
            width=5,
            textvariable=self.bins_var,
        )
        self.bins_spin.pack(side="left", padx=5)

        self.stats_label = ttk.Label(self.frame, text="Select a column to plot")
        self.stats_label.pack(anchor="w", padx=5)

        self.figure = Figure(figsize=(6, 3))
        self.ax = self.figure.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

        self.current_x: list[int] = []
        self.current_y: list[float | None] = []
        self.is_time_data = False

    def set_series(self, x_values: list[int], y_values: list[float | None], is_time: bool = False):
        self.current_x = x_values
        self.current_y = y_values
        self.is_time_data = is_time
        self.mode = "plot"
        self._update_mode_label()
        self._draw_plot()

    def _toggle_grid(self):
        self._draw_plot()

    def _toggle_mode(self):
        if not self.current_y:
            self._show_no_data()
            return
        self.mode = "hist" if self.mode == "plot" else "plot"
        self._update_mode_label()
        self._draw_plot()

    def _update_mode_label(self):
        self.toggle_btn.configure(text="Plot" if self.mode == "hist" else "Histogram")

    def _draw_plot(self):
        if not self.current_y:
            self._show_no_data()
            return
        valid_values = [v for v in self.current_y if v is not None]
        if not valid_values:
            self._show_no_data()
            return
        if self.mode == "hist":
            self.plot_hist(valid_values, is_time=self.is_time_data)
        else:
            self.plot_series(self.current_x, self.current_y, is_time=self.is_time_data)

    def plot_series(self, x_values: list[int], y_values: list[float | None], is_time: bool = False):
        self.ax.clear()

        points = [(x, y) for x, y in zip(x_values, y_values) if y is not None]
        if not points:
            self._show_no_data()
            return

        xs, ys = zip(*points)
        self.ax.scatter(xs, ys, color=self.color)

        for (prev_x, prev_y), (cur_x, cur_y) in zip(points, points[1:]):
            if prev_y is None or cur_y is None:
                continue
            style = "solid" if cur_x == prev_x + 1 else "dashed"
            self.ax.plot([prev_x, cur_x], [prev_y, cur_y], linestyle=style, color=self.color)

        if is_time:
            self.ax.yaxis.set_major_formatter(FuncFormatter(lambda _, pos: self._format_seconds_label(_)))

        self.ax.set_xlabel("Shot")
        self.ax.grid(self.grid_var.get())
        self.canvas.draw_idle()
        self._update_stats([v for v in y_values if v is not None], is_time=is_time)

    def plot_hist(self, values: list[float], is_time: bool = False):
        self.ax.clear()
        bins = self._safe_bins()
        self.ax.hist(values, bins=bins, color=self.color, edgecolor="black")
        self.ax.set_xlabel("Value")
        self.ax.set_ylabel("Frequency")
        if is_time:
            self.ax.xaxis.set_major_formatter(FuncFormatter(lambda _, pos: self._format_seconds_label(_)))
        self.ax.grid(self.grid_var.get())
        self.canvas.draw_idle()
        self._update_stats(values, is_time=is_time)

    def _safe_bins(self) -> int:
        try:
            bins = int(self.bins_var.get())
            return max(1, min(100, bins))
        except Exception:
            return 5

    def _update_stats(self, values: list[float], is_time: bool):
        if not values:
            self.stats_label.configure(text="No numeric data")
            return
        min_v = min(values)
        max_v = max(values)
        mean_v = mean(values)
        median_v = median(values)
        if is_time:
            formatter = self._format_seconds_label
            txt = (
                f"Min: {formatter(min_v)} | Max: {formatter(max_v)} | "
                f"Mean: {formatter(mean_v)} | Median: {formatter(median_v)}"
            )
        else:
            txt = (
                f"Min: {min_v:.3f} | Max: {max_v:.3f} | "
                f"Mean: {mean_v:.3f} | Median: {median_v:.3f}"
            )
        self.stats_label.configure(text=txt)

    def _show_no_data(self):
        self.ax.clear()
        self.ax.text(0.5, 0.5, "No numeric data", ha="center", va="center")
        self.ax.grid(self.grid_var.get())
        self.canvas.draw_idle()
        self.stats_label.configure(text="No numeric data")

    @staticmethod
    def _format_seconds_label(value):
        try:
            total_seconds = int(round(float(value)))
        except Exception:
            return "00:00:00"
        h = total_seconds // 3600
        m = (total_seconds % 3600) // 60
        s = total_seconds % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

# ==========================
# Main GUI
# ==========================

class ShotLogReader:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Shot Log Reader 3")

        self.log_path: Path | None = None
        self.manual_path: Path | None = None
        self.motor_path: Path | None = None

        self.log_analyzer = LogShotAnalyzer()
        self.manual_rows: list[list[str]] = []
        self.motor_rows: list[list[str]] = []
        self.manual_header: list[str] = []
        self.motor_header: list[str] = []

        self.shots_data: list[dict] = []
        self.camera_summary: list[dict] = []
        self.global_summary: dict = {}

        self.observer = Observer()
        self.watch_schedules = []
        self._handlers: list[TargetWatcher] = []

        self.previous_rows = {
            "log": [],
            "manual": [],
            "motor": [],
        }

        self.manual_view_rows: list[ShotViewRow] = []
        self.motor_view_rows: list[ShotViewRow] = []

        self._build_ui()

    # ---------- UI ----------
    def _build_ui(self):
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(fill="x", padx=10, pady=10)

        tk.Button(btn_frame, text="Select LOG (.txt)", command=self._select_log).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Select MANUAL CSV", command=self._select_manual).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Select MOTOR CSV", command=self._select_motor).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Export Excel", command=self._export_excel).pack(side="right", padx=5)

        frm_global = ttk.LabelFrame(self.root, text="Global summary")
        frm_global.pack(fill="x", padx=10, pady=5)
        self.lbl_global = ttk.Label(frm_global, text="No data")
        self.lbl_global.pack(anchor="w", padx=5, pady=5)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        log_frame = ttk.Frame(self.notebook)
        manual_frame = ttk.Frame(self.notebook)
        motor_frame = ttk.Frame(self.notebook)
        self.notebook.add(log_frame, text="Logs")
        self.notebook.add(manual_frame, text="Manual Params")
        self.notebook.add(motor_frame, text="Motor Params")

        self.manual_plot_panel = ColumnPlotPanel(manual_frame, color="blue")
        self.motor_plot_panel = ColumnPlotPanel(motor_frame, color="red")

        frm_cam = ttk.LabelFrame(log_frame, text="Per-camera summary")
        frm_cam.pack(fill="x", padx=5, pady=5)
        columns_cam = ("camera", "shots_used", "shots_missing")
        self.tree_cam = ttk.Treeview(frm_cam, columns=columns_cam, show="headings", height=5)
        for col in columns_cam:
            self.tree_cam.heading(col, text=col)
            self.tree_cam.column(col, width=140, anchor="center")
        self.tree_cam.pack(side="left", fill="x", expand=True)
        scrollbar_cam = ttk.Scrollbar(frm_cam, orient="vertical", command=self.tree_cam.yview)
        self.tree_cam.configure(yscrollcommand=scrollbar_cam.set)
        scrollbar_cam.pack(side="right", fill="y")
        self.tree_cam.tag_configure("cam_ok", background="green", foreground=TEXT_DEFAULT)
        self.tree_cam.tag_configure("cam_missing", background="red", foreground=TEXT_DEFAULT)

        frm_shots = ttk.LabelFrame(log_frame, text="Shots")
        frm_shots.pack(fill="both", expand=True, padx=5, pady=5)

        columns_shot = (
            "shot_number",
            "missing_count",
            "expected_count",
            "trigger_time",
            "min_time",
            "max_time",
            "missing_cams",
            "trigger_camera",
            "first_camera",
            "last_camera",
            "expected_cams",
            "trigger_cams",
        )
        headers = {
            "shot_number": "Shot #",
            "missing_count": "# Missing",
            "expected_count": "# Expected",
            "trigger_time": "Trigger time",
            "min_time": "Min time",
            "max_time": "Max time",
            "missing_cams": "Missing cameras",
            "trigger_camera": "Trigger camera",
            "first_camera": "First camera",
            "last_camera": "Last camera",
            "expected_cams": "Expected cameras",
            "trigger_cams": "Trigger cameras",
        }
        self.tree_shot = ttk.Treeview(frm_shots, columns=columns_shot, show="headings")
        for col in columns_shot:
            self.tree_shot.heading(col, text=headers[col])
            width = 260 if col in {"missing_cams", "expected_cams", "trigger_cams"} else 120
            self.tree_shot.column(col, width=width, anchor="center")

        self.tree_shot.pack(side="left", fill="both", expand=True)
        scrollbar_shot = ttk.Scrollbar(frm_shots, orient="vertical", command=self.tree_shot.yview)
        self.tree_shot.configure(yscrollcommand=scrollbar_shot.set)
        scrollbar_shot.pack(side="right", fill="y")
        self.tree_shot.tag_configure("ok", background="green", foreground=TEXT_DEFAULT)
        self.tree_shot.tag_configure("missing", background="red", foreground=TEXT_DEFAULT)

        self.csv_trees: dict[str, ttk.Treeview] = {}
        self.plot_panels: dict[str, ColumnPlotPanel] = {
            "manual": self.manual_plot_panel,
            "motor": self.motor_plot_panel,
        }
        for key, frame in ("manual", manual_frame), ("motor", motor_frame):
            tree_frame = ttk.Frame(frame)
            tree_frame.pack(fill="both", expand=True)
            tree = ttk.Treeview(tree_frame, columns=[], show="headings")
            tree.pack(fill="both", expand=True)
            tree.tag_configure("bg_green", background=GREEN_BG, foreground=TEXT_DEFAULT)
            tree.tag_configure("bg_blue", background=BLUE_BG, foreground=TEXT_DEFAULT)
            tree.tag_configure("bg_red", background=RED_BG, foreground=TEXT_DEFAULT)
            tree.tag_configure("bg_orange", background=ORANGE_BG, foreground=TEXT_DEFAULT)
            tree.tag_configure("fg_yellow", foreground=TEXT_WARNING)
            tree.bind("<Button-1>", lambda e, src=key: self._on_csv_heading_click(src, e))
            self.csv_trees[key] = tree

        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ---------- File selection ----------
    def _select_log(self):
        selected = filedialog.askopenfilename(title="Select log file", filetypes=[("Log files", "*.txt"), ("All", "*.*")])
        if not selected:
            return
        self.log_path = Path(selected)
        self._refresh_watchers()
        self._parse_all()

    def _select_manual(self):
        selected = filedialog.askopenfilename(title="Select manual CSV", filetypes=[("CSV", "*.csv"), ("All", "*.*")])
        if not selected:
            return
        self.manual_path = Path(selected)
        self._refresh_watchers()
        self._parse_all()

    def _select_motor(self):
        selected = filedialog.askopenfilename(title="Select motor CSV", filetypes=[("CSV", "*.csv"), ("All", "*.*")])
        if not selected:
            return
        self.motor_path = Path(selected)
        self._refresh_watchers()
        self._parse_all()

    # ---------- Watchdog ----------
    def _refresh_watchers(self):
        # Stop observer completely
        if self.observer.is_alive():
            self.observer.stop()
            self.observer.join()

        # Create a NEW observer
        self.observer = Observer()
        self.watch_schedules = []
        self._handlers = []

        for path in [self.log_path, self.manual_path, self.motor_path]:
            if path is None:
                continue
            watcher = TargetWatcher(path, self._parse_all, self.root)
            self._handlers.append(watcher)
            schedule = self.observer.schedule(watcher, path.parent, recursive=False)
            self.watch_schedules.append(schedule)

        self.observer.start()

    # ---------- Parsing ----------
    def _parse_all(self):
        try:
            if self.log_path:
                self.shots_data = self.log_analyzer.parse_log_file(self.log_path)
                self.camera_summary = self.log_analyzer.compute_camera_summary()
                self.global_summary = self.log_analyzer.compute_global_summary()
            else:
                self.shots_data = []
                self.camera_summary = []
                self.global_summary = {}
            if self.manual_path:
                self.manual_header, self.manual_rows = self._parse_csv(self.manual_path)
                self._configure_csv_tree("manual", self.manual_header)
            else:
                self.manual_header, self.manual_rows = [], []
                self._configure_csv_tree("manual", [])
            if self.motor_path:
                self.motor_header, self.motor_rows = self._parse_csv(self.motor_path)
                self._configure_csv_tree("motor", self.motor_header)
            else:
                self.motor_header, self.motor_rows = [], []
                self._configure_csv_tree("motor", [])
            self._refresh_views()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _parse_csv(self, path: Path) -> tuple[list[str], list[list[str]]]:
        header: list[str] = []
        rows: list[list[str]] = []
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f, delimiter=",", quotechar='"')
            header = next(reader, [])
            for row in reader:
                rows.append(row)
        return header, rows

    def _configure_csv_tree(self, source: str, header: list[str]):
        tree = self.csv_trees[source]
        tree.delete(*tree.get_children())
        tree["columns"] = header
        tree["show"] = "headings"
        for col in header:
            tree.heading(col, text=col)
            tree.column(col, width=140, anchor="center")

    def _on_csv_heading_click(self, source: str, event):
        tree = self.csv_trees.get(source)
        if tree is None:
            return
        if tree.identify_region(event.x, event.y) != "heading":
            return
        column_id = tree.identify_column(event.x)
        try:
            col_index = int(column_id.lstrip("#")) - 1
        except Exception:
            return
        header = self.manual_header if source == "manual" else self.motor_header
        if col_index < 0 or col_index >= len(header):
            return
        normalized = self._normalize_header(header[col_index])
        if normalized in {"shot", "shot_number", "shot_", "shot__", "index", "shot#"}:
            return

        rows = self.manual_view_rows if source == "manual" else self.motor_view_rows
        x_vals, y_vals, is_time = self._collect_series(header, rows, col_index)
        panel = self.plot_panels.get(source)
        if panel:
            panel.set_series(x_vals, y_vals, is_time=is_time)

    # ---------- Table rendering ----------
    def _refresh_views(self):
        self._refresh_global_summary()
        self._refresh_camera_table()
        self._refresh_shots_table()

        self.previous_rows = {"log": [], "manual": [], "motor": []}
        log_rows = self._build_log_rows()
        manual_rows = self._build_csv_rows(self.manual_header, self.manual_rows, "manual")
        motor_rows = self._build_csv_rows(self.motor_header, self.motor_rows, "motor")

        all_keys = {row.key for row in log_rows} | {row.key for row in manual_rows} | {row.key for row in motor_rows}
        yellow_keys = self._compute_yellow_keys(log_rows, manual_rows, motor_rows, all_keys)
        log_rows = self._apply_log_backgrounds(log_rows, yellow_keys)

        if self.manual_header:
            manual_rows = self._ensure_rows(manual_rows, all_keys, "manual", header=self.manual_header)
        if self.motor_header:
            motor_rows = self._ensure_rows(motor_rows, all_keys, "motor", header=self.motor_header)

        self.manual_view_rows = manual_rows
        self.motor_view_rows = motor_rows

        self._render_rows("manual", manual_rows, yellow_keys)
        self._render_rows("motor", motor_rows, yellow_keys)

    def _build_log_rows(self) -> list[ShotViewRow]:
        rows: list[ShotViewRow] = []
        for shot in sorted(self.shots_data, key=lambda s: (s.get("date", ""), s.get("shot_number", -1))):
            key = self._make_key(shot.get("shot_number"), self._format_time(shot.get("trigger_time")))
            values = [shot.get("date", ""), f"{shot.get('shot_number', 0):04d}"]
            rows.append(
                ShotViewRow(
                    key=key,
                    values=values,
                    bg="red" if shot.get("missing_cams") else "green",
                    yellow_text=False,
                    incomplete=bool(shot.get("missing_cams")),
                )
            )
        return rows

    def _format_time(self, dt):
        return dt.strftime("%H:%M:%S") if dt else ""

    def _refresh_global_summary(self):
        gs = self.global_summary
        if not gs or gs.get("total_shots", 0) == 0:
            self.lbl_global.configure(text="No shots found in log.")
            return

        dates = gs["dates"]
        date_txt = dates[0] if len(dates) == 1 else ", ".join(dates)

        def fmt_time(dt):
            return dt.strftime("%H:%M:%S") if dt else "-"

        txt = (
            f"Date(s): {date_txt} | "
            f"Start: {fmt_time(gs.get('start_time'))} | "
            f"End: {fmt_time(gs.get('end_time'))} | "
            f"Total shots: {gs.get('total_shots', 0)} | "
            f"Shots with missing cameras: {gs.get('shots_with_missing', 0)}"
        )
        self.lbl_global.configure(text=txt)

    def _refresh_camera_table(self):
        self.tree_cam.delete(*self.tree_cam.get_children())
        for row in self.camera_summary:
            cam = row["camera"]
            used = row["shots_used"]
            missing = row["shots_missing"]
            tag = "cam_ok" if missing == 0 else "cam_missing"
            self.tree_cam.insert("", "end", values=(cam, used, missing), tags=(tag,))

    def _refresh_shots_table(self):
        self.tree_shot.delete(*self.tree_shot.get_children())
        for shot in self.shots_data:
            missing_cams = sorted(shot["missing_cams"]) if shot["missing_cams"] else []
            expected_cams = sorted(shot["expected_cams"]) if shot["expected_cams"] else []
            trigger_cams = sorted(shot["trigger_cams"]) if shot["trigger_cams"] else []

            missing_str = "[" + ", ".join(missing_cams) + "]"
            expected_str = "[" + ", ".join(expected_cams) + "]"
            trigger_str = "[" + ", ".join(trigger_cams) + "]"

            missing_count = len(missing_cams)
            expected_count = len(expected_cams)

            trigger_cam = shot.get("trigger_camera") or ""
            first_cam = shot.get("first_camera") or ""
            last_cam = shot.get("last_camera") or ""

            values = (
                f"{shot['shot_number']:03d}",
                missing_count,
                expected_count,
                self._format_time(shot.get("trigger_time")),
                self._format_time(shot.get("min_time")),
                self._format_time(shot.get("max_time")),
                missing_str,
                trigger_cam,
                first_cam,
                last_cam,
                expected_str,
                trigger_str,
            )

            tag = "ok" if missing_count == 0 else "missing"
            self.tree_shot.insert("", "end", values=values, tags=(tag,))

    def _build_csv_rows(self, header: list[str], csv_rows: list[list[str]], source: str) -> list[ShotViewRow]:
        rows: list[ShotViewRow] = []
        if not header:
            return rows

        for csv_row in csv_rows:
            original_len = len(csv_row)
            values = csv_row + [""] * (len(header) - len(csv_row))
            key = self._extract_key_from_header(header, values)
            incomplete = self._is_csv_row_incomplete(header, values, original_len, source)
            rows.append(
                ShotViewRow(
                    key=key,
                    values=values,
                    bg="red" if incomplete else "blue",
                    yellow_text=False,
                    incomplete=incomplete,
                )
            )
        rows.sort(key=lambda r: (r.key[0], r.key[1]))
        prev_values: list[str] | None = None
        for row in rows:
            row.bg = self._determine_generic_bg(prev_values, row.values, row.incomplete)
            prev_values = row.values
        return rows

    def _ensure_rows(
        self,
        rows: list[ShotViewRow],
        all_keys: set[tuple[int, str]],
        source: str,
        header: list[str] | None = None,
    ) -> list[ShotViewRow]:
        existing = {r.key for r in rows}
        for key in all_keys - existing:
            shot_idx, trigger_time = key
            shot_disp = f"{shot_idx:04d}" if shot_idx and shot_idx > 0 else ""
            cols = header or []
            values = [""] * len(cols)
            if cols:
                shot_col = self._find_header_index(cols, {"shot", "shot_number", "shot_", "shot__", "index", "shot#"})
                time_col = self._find_header_index(cols, {"trigger_time", "time", "trigger_time_"})
                if shot_col is not None and shot_col < len(values):
                    values[shot_col] = shot_disp
                if time_col is not None and time_col < len(values):
                    values[time_col] = trigger_time
            bg = "red"
            rows.append(ShotViewRow(key=key, values=values, bg=bg, yellow_text=False, incomplete=True))
        rows.sort(key=lambda r: (r.key[0], r.key[1]))
        return rows

    def _determine_generic_bg(self, prev_values: list[str] | None, values: list[str], incomplete: bool) -> str:
        if incomplete:
            return "red"
        if prev_values is None or prev_values == values:
            return "blue"
        return "green"

    def _compute_yellow_keys(self, log_rows, manual_rows, motor_rows, all_keys):
        yellow = set()

        def build_counts(rows):
            counts = {}
            incomplete_keys = set()
            for r in rows:
                counts[r.key] = counts.get(r.key, 0) + 1
                if r.incomplete:
                    incomplete_keys.add(r.key)
            return counts, incomplete_keys

        manual_counts, manual_incomplete = build_counts(manual_rows)
        motor_counts, motor_incomplete = build_counts(motor_rows)

        for key in all_keys:
            manual_issue = manual_counts.get(key, 0) != 1 or key in manual_incomplete
            motor_issue = motor_counts.get(key, 0) != 1 or key in motor_incomplete
            if manual_issue or motor_issue:
                yellow.add(key)
        return yellow

    def _apply_log_backgrounds(self, log_rows: list[ShotViewRow], yellow_keys: set[tuple[int, str]]):
        prev_values: list[str] | None = None
        for row in log_rows:
            if row.incomplete:
                row.bg = "red"
            else:
                csv_ok = row.key not in yellow_keys
                if csv_ok and (prev_values is None or prev_values == row.values):
                    row.bg = "blue"
                else:
                    row.bg = "green"
            prev_values = row.values
        return log_rows

    @staticmethod
    def _make_key(shot_num: int | None, trigger_time: str | None) -> tuple[int, str]:
        try:
            shot_idx = int(shot_num) if shot_num is not None else -1
        except Exception:
            shot_idx = -1
        return shot_idx, trigger_time or ""

    def _extract_key_from_header(self, header: list[str], values: list[str]) -> tuple[int, str]:
        shot_idx = self._find_header_index(header, {"shot", "shot_number", "shot_", "shot__", "index", "shot#"})
        time_idx = self._find_header_index(header, {"trigger_time", "time", "trigger_time_"})
        shot_val = values[shot_idx] if shot_idx is not None and shot_idx < len(values) else ""
        trigger_time = values[time_idx] if time_idx is not None and time_idx < len(values) else ""
        return self._make_key(shot_val if shot_val != "" else -1, trigger_time)

    def _is_csv_row_incomplete(self, header: list[str], values: list[str], original_len: int, source: str) -> bool:
        shot_idx = self._find_header_index(header, {"shot", "shot_number", "shot_", "shot__", "index", "shot#"})
        time_idx = self._find_header_index(header, {"trigger_time", "time", "trigger_time_"})
        shot_val = values[shot_idx].strip() if shot_idx is not None and shot_idx < len(values) else ""
        time_val = values[time_idx].strip() if time_idx is not None and time_idx < len(values) else ""
        shot_valid = self._parse_int_or_none(shot_val) is not None
        missing_fields = original_len < len(header)
        if source == "manual":
            return missing_fields or not shot_valid or time_val == ""

        other_values = [
            values[i].strip()
            for i in range(len(header))
            if i < len(values) and i not in {shot_idx, time_idx}
        ]
        empty_other = all(v == "" for v in other_values)
        return missing_fields or shot_val == "" or time_val == "" or empty_other

    @staticmethod
    def _normalize_header(text: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", text.strip().lower())

    def _find_header_index(self, header: list[str], names: set[str]) -> int | None:
        normalized = [self._normalize_header(h) for h in header]
        for i, name in enumerate(normalized):
            if name in names:
                return i
        return None

    @staticmethod
    def _parse_int_or_none(value: str) -> int | None:
        try:
            return int(value)
        except Exception:
            return None

    @staticmethod
    def _parse_float_or_none(value: str) -> float | None:
        try:
            return float(value)
        except Exception:
            return None

    @staticmethod
    def _parse_time_to_seconds(value: str) -> float | None:
        try:
            t = datetime.strptime(value.strip(), "%H:%M:%S")
            return float(t.hour * 3600 + t.minute * 60 + t.second)
        except Exception:
            return None

    def _collect_series(self, header: list[str], rows: list[ShotViewRow], value_idx: int):
        shot_idx = self._find_header_index(header, {"shot", "shot_number", "shot_", "shot__", "index", "shot#"})
        normalized_value = self._normalize_header(header[value_idx]) if value_idx < len(header) else ""
        is_time = normalized_value in {"trigger_time", "time", "trigger_time_"}

        x_values: list[int] = []
        y_values: list[float | None] = []
        for row in rows:
            shot_val = None
            if shot_idx is not None and shot_idx < len(row.values):
                shot_val = self._parse_int_or_none(row.values[shot_idx].strip())
            if shot_val is None or shot_val < 0:
                shot_val = row.key[0]
            if shot_val is None or shot_val < 0:
                continue

            raw_value = row.values[value_idx] if value_idx < len(row.values) else ""
            if is_time:
                parsed_value = self._parse_time_to_seconds(raw_value) if raw_value else None
            else:
                parsed_value = self._parse_float_or_none(raw_value) if raw_value != "" else None

            x_values.append(shot_val)
            y_values.append(parsed_value)

        return x_values, y_values, is_time

    def _render_rows(self, source: str, rows: list[ShotViewRow], yellow_keys: set[tuple[int, str]]):
        tree = self.csv_trees[source]
        tree.delete(*tree.get_children())
        for row in rows:
            tags = []
            if row.bg == "green":
                tags.append("bg_green")
            elif row.bg == "blue":
                tags.append("bg_blue")
            elif row.bg == "red":
                tags.append("bg_red")
            elif row.bg == "orange":
                tags.append("bg_orange")
            if row.key in yellow_keys or row.yellow_text:
                tags.append("fg_yellow")
            tree.insert("", "end", values=row.values, tags=tags)

    # ---------- Excel export ----------
    def _export_excel(self):
        if not any([self.log_path, self.manual_path, self.motor_path]):
            messagebox.showerror("Error", "Select at least one source to export")
            return
        xlsx_path = filedialog.asksaveasfilename(
            title="Export to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not xlsx_path:
            return
        log_rows = self._build_log_rows()
        manual_rows = self._build_csv_rows(self.manual_header, self.manual_rows, "manual")
        motor_rows = self._build_csv_rows(self.motor_header, self.motor_rows, "motor")
        all_keys = {row.key for row in log_rows} | {row.key for row in manual_rows} | {row.key for row in motor_rows}
        yellow_keys = self._compute_yellow_keys(log_rows, manual_rows, motor_rows, all_keys)
        log_rows = self._apply_log_backgrounds(log_rows, yellow_keys)

        wb = Workbook()

        ws_logs = wb.active
        ws_logs.title = "Logs"
        log_headers = [
            "Shot #",
            "# Missing",
            "# Expected",
            "Trigger time",
            "Min time",
            "Max time",
            "Missing cameras",
            "Trigger camera",
            "First camera",
            "Last camera",
            "Expected cameras",
            "Trigger cameras",
        ]
        ws_logs.append(log_headers)

        shot_rows_excel: list[ShotViewRow] = []
        for shot in self.shots_data:
            missing_cams = sorted(shot["missing_cams"]) if shot["missing_cams"] else []
            expected_cams = sorted(shot["expected_cams"]) if shot["expected_cams"] else []
            trigger_cams = sorted(shot["trigger_cams"]) if shot["trigger_cams"] else []

            missing_str = "[" + ", ".join(missing_cams) + "]"
            expected_str = "[" + ", ".join(expected_cams) + "]"
            trigger_str = "[" + ", ".join(trigger_cams) + "]"

            missing_count = len(missing_cams)
            expected_count = len(expected_cams)

            trigger_cam = shot.get("trigger_camera") or ""
            first_cam = shot.get("first_camera") or ""
            last_cam = shot.get("last_camera") or ""

            values = [
                f"{shot['shot_number']:03d}",
                missing_count,
                expected_count,
                self._format_time(shot.get("trigger_time")),
                self._format_time(shot.get("min_time")),
                self._format_time(shot.get("max_time")),
                missing_str,
                trigger_cam,
                first_cam,
                last_cam,
                expected_str,
                trigger_str,
            ]

            row_obj = ShotViewRow(
                key=self._make_key(shot.get("shot_number"), self._format_time(shot.get("trigger_time"))),
                values=values,
                bg="green" if missing_count == 0 else "red",
                yellow_text=False,
                incomplete=missing_count != 0,
            )
            shot_rows_excel.append(row_obj)
            ws_logs.append(values)
            self._apply_excel_styles(ws_logs, ws_logs.max_row, row_obj, set())

        ws_cam = wb.create_sheet("Per_camera_summary")
        ws_cam.append(["Camera", "Shots requested", "Shots missing for this camera"])
        for row in self.camera_summary:
            cam = row["camera"]
            used = row["shots_used"]
            missing = row["shots_missing"]
            ws_cam.append([cam, used, missing])
            excel_row = ws_cam.max_row
            fill_color = GREEN_BG.lstrip("#") if missing == 0 else RED_BG.lstrip("#")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            font = Font(color=to_argb(TEXT_DEFAULT))
            for cell in ws_cam[excel_row]:
                cell.fill = fill
                cell.font = font

        ws_manual = wb.create_sheet("Manual_Params")
        ws_motor = wb.create_sheet("Motor_Params")
        for ws, rows, headers in (
            (ws_manual, manual_rows, self.manual_header),
            (ws_motor, motor_rows, self.motor_header),
        ):
            ws.append(headers)
            for row in rows:
                ws.append(row.values)
                self._apply_excel_styles(ws, ws.max_row, row, yellow_keys)

        try:
            wb.save(xlsx_path)
            messagebox.showinfo("Export", f"Exported to {xlsx_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}")

    def _apply_excel_styles(self, ws, row_idx: int, row: ShotViewRow, yellow_keys):
        color_map = {
            "green": GREEN_BG.lstrip("#"),
            "blue": BLUE_BG.lstrip("#"),
            "red": RED_BG.lstrip("#"),
            "orange": ORANGE_BG.lstrip("#"),
        }
        fill_color = color_map.get(row.bg)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
        font_color = TEXT_WARNING.lstrip("#") if (row.key in yellow_keys or row.yellow_text) else TEXT_DEFAULT
        font = Font(color=to_argb(font_color))
        for cell in ws[row_idx]:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font

    # ---------- Close ----------
    def _on_close(self):
        try:
            if self.observer.is_alive():
                self.observer.stop()
                self.observer.join(timeout=1)
        finally:
            self.root.destroy()


def main():
    root = tk.Tk()
    app = ShotLogReader(root)
    root.mainloop()


if __name__ == "__main__":
    main()
